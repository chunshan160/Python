from openpyxl import load_workbook


class ReadExcel:
    """封装读取excel文件及写回数据的方法"""

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        """打开工作簿，选择表单"""
        self.workbook = load_workbook(self.file_name)
        self.sheet = self.workbook[self.sheet_name]

    def close(self):
        """关闭工作簿"""
        self.workbook.close()

    def get_header(self):
        """获取表头"""
        self.open()
        title_list = []
        for i in range(1, self.sheet.max_column+1):
            title_list.append(self.sheet.cell(1, i).value)
        return title_list

    def get_data(self):
        """获取测试数据"""
        self.open()
        test_data = []
        for item in range(2, self.sheet.max_row+1):
            case_data = {}
            case_data['case_id'] = self.sheet.cell(item, 1).value
            case_data['interface'] = self.sheet.cell(item, 2).value
            case_data['title'] = self.sheet.cell(item, 3).value
            case_data['url'] = self.sheet.cell(item, 4).value
            case_data['data'] = self.sheet.cell(item, 5).value
            case_data['method'] = self.sheet.cell(item, 6).value
            case_data['expected'] = self.sheet.cell(item, 7).value
            test_data.append(case_data)
        self.close()
        return test_data

    def write_data(self, item, result):
        """
        写回测试结果数据
        :param item: xlsx文件中的行数，必须大于1
        :param result: 测试结果
        """
        self.open()
        self.sheet.cell(item, 8).value = result
        self.workbook.save(self.file_name)
        self.close()


if __name__ == '__main__':
    from interface_auto.common.project_path import data_dir
    import os
    test_data = os.path.join(data_dir, 'data.xlsx')
    res1 = ReadExcel(test_data, 'login').get_header()
    print(res1)
    res2 = ReadExcel(test_data, 'register').get_data()
    for data in res2:
        print(data)
    # ReadExcel(test_data, 'login').write_data(2, '通过')
