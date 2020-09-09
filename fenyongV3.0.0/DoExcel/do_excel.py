#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/9 9:55
# @Author :春衫
# @File :do_excel.py

from tools.project_path import *
from openpyxl import load_workbook
from tools.read_config import ReadConfig
from DoMysql.sql import SQL


class DoExcel:

    @classmethod
    def get_data(cls, file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))  # 配置文件的内容 字典

        test_data = []
        for key in mode:  # 遍历这个存在配置文件里面的字典 也就是表单名
            sheet = wb[key]  # 打开Excel里的这个表单
            # 跑全部用例
            if mode[key]["焕商分佣"] == 'all':  # 判断value
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['title'] = sheet.cell(i, 2).value
                    row_data['member_level'] = sheet.cell(i, 3).value
                    row_data['buyer_identity'] = sheet.cell(i, 4).value
                    row_data['seller_identity'] = sheet.cell(i, 5).value
                    row_data['payment_method'] = sheet.cell(i, 6).value
                    row_data['goodsname'] = sheet.cell(i + 1, 7).value
                    row_data['data'] = sheet.cell(i, 8).value
                    row_data['operational_setting'] = sheet.cell(i, 9).value
                    row_data['proportion'] = sheet.cell(i, 10).value
                    row_data['superior'] = sheet.cell(i, 11).value
                    row_data['order'] = sheet.cell(i, 12).value
                    row_data['reserve_fund'] = sheet.cell(i, 13).value
                    row_data['bind_relationship_data'] = sheet.cell(i, 14).value
                    row_data['second_payagent_ratio'] = sheet.cell(i, 15).value
                    row_data['sheet_name'] = key
                    row_data['surroundings'] = mode[key]["环境"]
                    row_data['payPassword']=mode[key]["支付密码"]
                    test_data.append(row_data)
            # 跑指定用例
            else:
                for i in mode[key]["焕商分佣"]:  # case_id
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i + 1, 1).value
                    row_data['title'] = sheet.cell(i + 1, 2).value
                    row_data['member_level'] = sheet.cell(i + 1, 3).value
                    row_data['buyer_identity'] = sheet.cell(i + 1, 4).value
                    row_data['seller_identity'] = sheet.cell(i + 1, 5).value
                    row_data['payment_method'] = sheet.cell(i + 1, 6).value
                    row_data['goodsname'] = sheet.cell(i + 1, 7).value
                    row_data['data'] = sheet.cell(i + 1, 8).value
                    row_data['operational_setting'] = sheet.cell(i + 1, 9).value
                    row_data['proportion'] = sheet.cell(i + 1, 10).value
                    row_data['superior'] = sheet.cell(i + 1, 11).value
                    row_data['order'] = sheet.cell(i + 1, 12).value
                    row_data['reserve_fund'] = sheet.cell(i + 1, 13).value
                    row_data['bind_relationship_data'] = sheet.cell(i + 1, 14).value
                    row_data['second_payagent_ratio'] = sheet.cell(i + 1, 15).value
                    row_data['sheet_name'] = key
                    row_data['surroundings'] = mode[key]["环境"]
                    row_data['payPassword'] = mode[key]["支付密码"]
                    test_data.append(row_data)

        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, actual_userid, actual_changes,
                   expected_result, actual_result, TestResult, Error):  # 专门写回数据  i行号  TestResult结果
        '''

        :param file_name: 文件名
        :param sheet_name: 表单名
        :param i: 用例id
        :param actual_userid: 实际用户id
        :param actual_changes: 实际变化金额
        :param expected_result: 预期结果
        :param actual_result: 实际结果
        :param TestResult: 测试结果
        :param Error: 错误原因
        :return:
        '''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]  # 选择表单
        sheet.cell(i, 16).value = actual_userid  # 实际
        sheet.cell(i, 17).value = actual_changes  # 实际
        sheet.cell(i, 18).value = expected_result  # 预期结果
        sheet.cell(i, 19).value = actual_result  # 实际结果
        sheet.cell(i, 20).value = TestResult  # 测试结果
        sheet.cell(i, 21).value = Error  # 报错内容
        wb.save(file_name)  # 保存结果

    @classmethod
    def fenyong_bili(cls, file_name, sheet_name, case_id, proportion):
        '''

                :param file_name: 文件名
                :param sheet_name: 表单名
                :param case_id: 用例id
                :param proportion: 省市区、个人分佣比例
                :return: 省市区、个人分佣比例 写回Excel
                '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 10).value = proportion
        wb.save(file_name)

    @classmethod
    def superior(cls, file_name, sheet_name, case_id, superior):
        '''

        :param file_name: 文件名
        :param sheet_name: 表单名
        :param case_id: 用例id
        :param superior:上级代理商/城市焕商
        :return:上级代理商/城市焕商写回Excel
        '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 11).value = superior
        wb.save(file_name)

    @classmethod
    def get_order(cls, file_name, sheet_name, case_id, order):
        '''

        :param file_name: 文件名
        :param sheet_name: 表单名
        :param case_id: 用例id
        :param order: 订单号
        :return: 订单id写回Excel
        '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 12).value = order
        wb.save(file_name)

    @classmethod
    def write_back_reserve_fund(cls, file_name, sheet_name, case_id, reserve_fund):
        '''

        :param file_name: 文件名
        :param sheet_name: 表单名
        :param case_id: 用例id
        :param reserve_fund: 未消耗充值金额和储备池
        :return: 未消耗充值金额和储备池 写回Excel
        '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 13).value = reserve_fund
        wb.save(file_name)

    @classmethod
    def bing_sale_id(cls, file_name, sheet_name, case_id, shangji_sale_id):
        '''
        :param file_name: 文件名
        :param sheet_name: 表单名
        :param case_id: 用例id
        :param shangji_sale_id: 买家上级销售/业务焕商/TCO/销售的上级
        :return: 买家上级销售/业务焕商/TCO/销售的上级 写回Excel
        '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 14).value = shangji_sale_id  # 计算
        wb.save(file_name)

    # 二级分佣比例
    def second_payagent_ratio(self, file_name, sheet_name, case_id, bili):

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 15).value = bili  # 计算
        wb.save(file_name)


if __name__ == '__main__':
    test_data = DoExcel.get_data(test_case_path)
    print(test_data)
    print(type(test_data))

    # test_data = DoExcel.get_order(test_case_path, 'Sheet2', 1000656, 1)
    # print(test_data)
