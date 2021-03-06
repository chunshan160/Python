#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/9 9:55
# @Author :春衫
# @File :do_excel.py

from tools.project_path import *
from openpyxl import load_workbook
from tools.read_config import ReadConfig
from Do_mysql.sql import SQL


class DoExcel:

    @classmethod
    def get_data(cls, file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))  # 配置文件的内容 字典

        test_data = []
        for key in mode:  # 遍历这个存在配置文件里面的字典 也就是表单名
            sheet = wb[key]  # 打开Excel里的这个表单
            # 跑全部用例
            if mode[key]["焕商分佣"] == 'all':  # 判断value
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['title'] = sheet.cell(i, 2).value
                    row_data['member_level'] = sheet.cell(i, 3).value
                    row_data['buyer_identity'] = sheet.cell(i, 4).value
                    row_data['seller_identity'] = sheet.cell(i, 5).value
                    row_data['payment_method'] = sheet.cell(i, 6).value
                    row_data['goodsname'] = sheet.cell(i + 1, 7).value
                    row_data['data'] = sheet.cell(i, 8).value
                    row_data['proportion'] = sheet.cell(i, 9).value
                    row_data['superior'] = sheet.cell(i, 10).value
                    row_data['order'] = sheet.cell(i, 11).value
                    row_data['reserve_fund'] = sheet.cell(i, 12).value
                    row_data['sheet_name'] = key
                    row_data['surroundings'] = mode[key]["环境"]
                    test_data.append(row_data)
            # 跑指定用例
            else:
                for i in mode[key]["焕商分佣"]:  # case_id
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i + 1, 1).value
                    row_data['title'] = sheet.cell(i + 1, 2).value
                    row_data['member_level'] = sheet.cell(i + 1, 3).value
                    row_data['buyer_identity'] = sheet.cell(i + 1, 4).value
                    row_data['seller_identity'] = sheet.cell(i + 1, 5).value
                    row_data['payment_method'] = sheet.cell(i + 1, 6).value
                    row_data['goodsname'] = sheet.cell(i + 1, 7).value
                    row_data['data'] = sheet.cell(i + 1, 8).value
                    row_data['proportion'] = sheet.cell(i + 1, 9).value
                    row_data['superior'] = sheet.cell(i + 1, 10).value
                    row_data['order'] = sheet.cell(i + 1, 11).value
                    row_data['reserve_fund'] = sheet.cell(i + 1, 12).value
                    row_data['sheet_name'] = key
                    row_data['surroundings'] = mode[key]["环境"]
                    test_data.append(row_data)

        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, expected_userid, actual_userid, expected_changes, actual_changes,
                   expected_result, actual_result, TestResult, Error):  # 专门写回数据  i行号  TestResult结果
        '''

        :param file_name: 文件名
        :param sheet_name: 表单名
        :param i: 用例id
        :param expected_userid: 预期用户id
        :param actual_userid: 实际用户id
        :param expected_changes: 预期变化金额
        :param actual_changes: 实际变化金额
        :param expected_result: 预期结果
        :param actual_result: 实际结果
        :param TestResult: 测试结果
        :param Error: 错误原因
        :return:
        '''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]  # 选择表单
        sheet.cell(i, 13).value = expected_userid  # 计算
        sheet.cell(i, 14).value = actual_userid  # 实际
        sheet.cell(i, 15).value = expected_changes  # 计算
        sheet.cell(i, 16).value = actual_changes  # 实际
        sheet.cell(i, 17).value = expected_result  # 预期结果
        sheet.cell(i, 18).value = actual_result  # 实际结果
        sheet.cell(i, 19).value = TestResult  # 测试结果
        sheet.cell(i, 20).value = Error  # 报错内容
        wb.save(file_name)  # 保存结果

    @classmethod
    def get_order(cls, ip, file_name, sheet_name, case_id, buyerid):
        '''

        :param file_name: 文件名
        :param sheet_name: 表单名
        :param case_id: 用例id
        :param buyerid: 买家id
        :return: 订单id写回Excel
        '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        # 查库获取最新的1条订单，一般是前端操作完，然后查库，获取订单
        order_data = SQL(ip).order(buyerid)
        # print(order_data)
        order = (order_data[0])[0]  # 返回是这种 ('EC-2020050611310900010204',)
        sheet.cell(case_id + 1, 11).value = order
        wb.save(file_name)
        return order

    @classmethod
    def superior(cls, file_name, sheet_name, case_id, superior):
        '''

        :param file_name: 文件名
        :param sheet_name: 表单名
        :param case_id: 用例id
        :param superior:上级代理商/城市焕商
        :return:上级代理商/城市焕商写回Excel
        '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 10).value = superior
        wb.save(file_name)

    @classmethod
    def reserve_fund(cls, file_name, sheet_name, case_id, reserve_fund):
        '''

        :param file_name: 文件名
        :param sheet_name: 表单名
        :param case_id: 用例id
        :param reserve_fund: 未消耗充值金额和储备池
        :return: 未消耗充值金额和储备池 写回Excel
        '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 12).value = reserve_fund
        wb.save(file_name)

    @classmethod
    def fenyong_bili(cls, file_name, sheet_name, case_id, fenyong_bili):
        '''

                :param file_name: 文件名
                :param sheet_name: 表单名
                :param case_id: 用例id
                :param fenyong_bili: 省市区、个人分佣比例
                :return: 省市区、个人分佣比例 写回Excel
                '''

        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(case_id + 1, 9).value = fenyong_bili
        wb.save(file_name)


if __name__ == '__main__':
    test_data = DoExcel.get_data(test_case_path)
    print(test_data)

    # test_data = DoExcel.get_order(test_case_path, 'Sheet2', 1000656, 1)
    # print(test_data)
