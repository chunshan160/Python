#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/9/5 16:09
# @Author :春衫
# @File :第09节_作业.py

from openpyxl import load_workbook

# # 1、转换成字典格式
# cook_str = 'RIDUPSID=D0727533D7147B7;PSTM=1530348042;' \
#            'BAIDUID=B1005c9BC2EB28;sugstore=0;__cfduid=d0a13458f8ac2a;' \
#            'BD_UPN=12314353;ispeed_lsm=2;BDORZ=B490B5EBF6F3CD402'
#
# dict = {i.split('=')[0]: i.split('=')[1] for i in cook_str.split(';')}
# print(dict)

# 2、当前有文件case.excel，.设计程序将excel中的用例读取到一个生成器?
def read_excel(file_path,sheet_name):
    """
    读取excle数据的方法
    :param file_path:文件路径
    :param sheet_name:表名
    :return:读取的数据
    """
    wb = load_workbook(file_path)
    ws=wb[sheet_name]
    row=ws.max_row +1#最大行数
    col=ws.max_column +1#最大列数
    test_data=({ws.cell(1,j).value:ws.cell(i,j).value for j in range(1,col)}for i in range(2,row))
    return test_data

res=read_excel('fenyong_data.xlsx', '焕商分佣_test')
for i in res:
    print(i)
