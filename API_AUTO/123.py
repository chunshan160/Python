#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/16 13:15
# @Author :春衫
# @File :123.py

from tools.project_path import *
from openpyxl import load_workbook
from tools.read_config import ReadConfig

# wb = load_workbook(self.file_name)
# sheet = wb[self.sheet_name]

# wb = load_workbook(test_case_path)
# sheet = wb["register"]
# phone_number = eval(sheet.cell(2, 4).value)
# print(phone_number)
# print((type(phone_number)))
# a = eval(phone_number["mobilephone"])
# print(a)
# print(type(a))
# b = a + 1
# print(b)
# phone_number["mobilephone"] = b
# print(phone_number)
# sheet.cell(2, 4).value = str(phone_number)
# wb.save(test_case_path)  # 保存结果

# phone_number = {'mobilephone': "${tel_1}", 'pwd': '123456', 'regname': 'ceshi'}
# a = eval(phone_number["mobilephone"])
# # b = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))
# print(a)
# print(type(a))
# b = a + 1
# phone_number["mobilephone"] = b
# print(phone_number)
# print(type(phone_number))



wb = load_workbook(test_case_path)
# mode = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))
sheet = wb["register"]
# case_id=1
a=sheet.max_row + 1
print(a)

# # for i in range(2, sheet.max_row + 1):
# if sheet.cell(3, 4).value.find('${tel_1}') != -1:
#     print("lalalala")