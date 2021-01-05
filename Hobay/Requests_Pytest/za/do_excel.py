#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/9 9:55
# @Author :春衫
# @File :do_excel.py

from Requests.tools.project_path import *
from openpyxl import load_workbook
from Requests.za.read_config2 import ReadConfig
from Requests.tools.get_data import GetData


class DoExcel:

    @classmethod
    def get_data(cls, file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))  # 配置文件的内容 字典

        tel = getattr(GetData, 'NoRegTel')  # 从Excel里面拿到的手机号 int
        test_data = []
        for key in mode:  # 遍历这个存在配置文件里面的字典 也就是表单名
            sheet = wb[key]  # 打开Excel里的这个表单
            if mode[key] == 'all':  # 判断value
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['title'] = sheet.cell(i, 2).value
                    row_data['url'] = sheet.cell(i, 3).value
                    # 做手机号的替换 正常登录用
                    if sheet.cell(i, 4).value.find('${tel}') != -1:  # -1是不存在 其他是存在该字符串
                        row_data['data'] = sheet.cell(i, 4).value.replace("${tel}", str(tel))  # 替换
                        cls.updata_tel(tel + 1, file_name, 'init')  # 更新手机号

                    # 加标
                    elif sheet.cell(i, 4).value.find('${admin_tel}') != -1:  # -1 存在该字符串
                        row_data['data'] = sheet.cell(i, 4).value.replace("${admin_tel}",
                                                                          str(getattr(GetData, 'admin_tel')))  # 替换
                    elif sheet.cell(i, 4).value.find('${loan_member_id}') != -1:  # -1 存在该字符串
                        row_data['data'] = sheet.cell(i, 4).value.replace("${loan_member_id}",
                                                                          str(getattr(GetData, 'loan_member_id')))  # 替换
                    elif sheet.cell(i, 4).value.find('${normal_tel}') != -1:  # -1 存在该字符串
                        row_data['data'] = sheet.cell(i, 4).value.replace("${normal_tel}",
                                                                          str(getattr(GetData, 'normal_tel')))  # 替换
                    elif sheet.cell(i, 4).value.find('${memberId}') != -1:  # -1 存在该字符串
                        row_data['data'] = sheet.cell(i, 4).value.replace("${memberId}",
                                                                          str(getattr(GetData,
                                                                                      'memberId')))  # 替换
                    else:
                        row_data['data'] = sheet.cell(i, 4).value
                    row_data['http_method'] = sheet.cell(i, 5).value
                    row_data['expected_code'] = str(sheet.cell(i, 6).value)
                    row_data['sheet_name'] = key
                    test_data.append(row_data)

            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id + 1, 1).value
                    row_data['title'] = sheet.cell(case_id + 1, 2).value
                    row_data['url'] = sheet.cell(case_id + 1, 3).value
                    # 做手机号的替换 正常登录用
                    if sheet.cell(case_id + 1, 4).value.find('${tel}') != -1:  # -1是不存在 其他是存在该字符串
                        row_data['data'] = sheet.cell(case_id + 1, 4).value.replace("${tel}", str(tel))  # 替换
                        cls.updata_tel(tel + 1, file_name, 'init')  # 更新手机号
                    # 加标
                    elif sheet.cell(case_id + 1, 4).value.find('${admin_tel}') != -1:  # -1 存在该字符串
                        row_data['data'] = sheet.cell(case_id + 1, 4).value.replace("${admin_tel}",
                                                                                    str(getattr(GetData,
                                                                                                'admin_tel')))  # 替换
                    elif sheet.cell(case_id + 1, 4).value.find('${loan_member_id}') != -1:  # -1 存在该字符串
                        row_data['data'] = sheet.cell(case_id + 1, 4).value.replace("${loan_member_id}",
                                                                                    str(getattr(GetData,
                                                                                                'loan_member_id')))  # 替换
                    elif sheet.cell(case_id + 1, 4).value.find('${normal_tel}') != -1:  # -1 存在该字符串
                        row_data['data'] = sheet.cell(case_id + 1, 4).value.replace("${normal_tel}",
                                                                                    str(getattr(GetData,
                                                                                                'normal_tel')))  # 替换
                    elif sheet.cell(case_id + 1, 4).value.find('${memberId}') != -1:  # -1 存在该字符串
                        row_data['data'] = sheet.cell(case_id + 1, 4).value.replace("${memberId}",
                                                                                    str(getattr(GetData,
                                                                                                'memberId')))  # 替换
                    else:
                        row_data['data'] = sheet.cell(case_id + 1, 4).value
                    row_data['http_method'] = sheet.cell(case_id + 1, 5).value
                    row_data['expected_code'] = str(sheet.cell(case_id + 1, 6).value)
                    row_data['sheet_name'] = key
                    test_data.append(row_data)

        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):  # 专门写回数据  i行号  result结果
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]  # 选择表单
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)  # 保存结果

    @classmethod
    def updata_tel(cls, tel, file_name, sheet_name):
        '''更新Excel表单里的手机号'''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(2, 1).value = tel
        wb.save(file_name)


if __name__ == '__main__':
    test_data = DoExcel.get_data(test_case_path)
    print(test_data)
