#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/9 9:55
# @Author :春衫
# @File :do_excel.py


from openpyxl import load_workbook

#方法二  需要用的时候读取所有的数据   就是磁盘读写要求高点

class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]
        self.max_row = self.sheet_obj.max_row
        # 获取一个表单对象

    def get_data(self, i, j):
        '''根据传入的坐标来获取值'''
        return self.sheet_obj.cell(i, j).value


if __name__ == '__main__':
    res = DoExcel("test.xlsx", "python").get_data(2, 1)
    print(res)
