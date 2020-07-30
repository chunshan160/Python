#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/9 9:55
# @Author :春衫
# @File :do_excel.py


from openpyxl import load_workbook


# 方法三

class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_header(self):
        '''获取第一行的标题行'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = []  # 存储我们的标题行
        for j in range(1, sheet.max_column + 1):  # 遍历列
            header.append(sheet.cell(1, j))

        return header

    def get_data(self):
        '''根据嵌套循环读取数据'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        header = self.get_header()
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            for j in range(1, sheet.max_column + 1):
                sub_data[header[j - 1]] = sheet.cell(i, j).value
            test_data.append(sub_data)
        return test_data


if __name__ == '__main__':
    print(DoExcel("test.xlsx", "python").get_data())
