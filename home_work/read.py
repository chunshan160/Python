#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/10/20 17:13
# @Author :春衫
# @File :read.py
from openpyxl import load_workbook


def get_data( file_name,sheet_name,name,type):
    wb = load_workbook(file_name)
    X_data = []
    Y_data = []
    sheet = wb[sheet_name]  # 打开Excel里的这个表单

    for a in range(1, sheet.max_column):
        type1 = sheet.cell(1, a).value
        for i in range(2, sheet.max_row + 1):
                name1 = sheet.cell(i, 1).value
                if name1 ==name or type1==type:
                        name2 = sheet.cell(i, 1).value
                        type2= sheet.cell(i, a).value
                        X_data.append(name2)
                        Y_data.append(type2)

    return X_data,Y_data

if __name__ == '__main__':
    a=get_data("D:\Pycharm_workspace\home_work\data.xlsx", "ceshi", None, "修复中")
    print(a)