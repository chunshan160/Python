#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/9 9:55
# @Author :春衫
# @File :do_excel.py


from openpyxl import load_workbook

#方法一 一次性读取所有的数据  对内存的要求会高点

class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data['test_case'] = sheet.cell(i, 1).value
            sub_data['url'] = sheet.cell(i, 2).value
            sub_data['data'] = sheet.cell(i, 3).value
            sub_data['method'] = sheet.cell(i, 4).value
            sub_data['expected_status'] = str(sheet.cell(i, 5).value)
            sub_data['expected_code'] = str(sheet.cell(i, 6).value)
            sub_data['expected_msg'] = sheet.cell(i, 7).value
            test_data.append(sub_data)
        return test_data


if __name__ == '__main__':
    print(DoExcel("test.xlsx", "python").get_data())
