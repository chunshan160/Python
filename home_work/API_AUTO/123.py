#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/16 13:15
# @Author :春衫
# @File :PDF加水印.py

from tools.project_path import *
from openpyxl import load_workbook
from tools.read_config import ReadConfig

# wb = load_workbook(self.file_name)
# sheet = wb[self.sheet_name]

# wb = load_workbook(test_case_path)
# sheet = wb["register"]
# phone_number = eval(sheet.cell(2, 4).value)
# print(phone_number)
# print((type(phone_number)))
# a = eval(phone_number["mobilephone"])
# print(a)
# print(type(a))
# b = a + 1
# print(b)
# phone_number["mobilephone"] = b
# print(phone_number)
# sheet.cell(2, 4).value = str(phone_number)
# wb.save(test_case_path)  # 保存结果

phone_number = {'mobilephone': 13724765586, 'pwd': '123456', 'regname': 'ceshi'}
a = eval(phone_number["mobilephone"])
# b = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))
print(a)
print(type(a))
b = a + 1
phone_number["mobilephone"] = b
print(phone_number)
print(type(phone_number))
