#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/9 9:55
# @Author :春衫
# @File :do_excel.py

from tools.project_path import *
from openpyxl import load_workbook
from tools.read_config import ReadConfig


class DoExcel:

    @staticmethod
    def get_data(file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))  # 配置文件的内容 字典
        test_data = []
        for key in mode:  # 遍历这个存在配置文件里面的字典 也就是表单名
            sheet = wb[key]  # 打开Excel里的这个表单
            if mode[key] == 'all':  # 判断value
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['title'] = sheet.cell(i, 2).value
                    row_data['url'] = sheet.cell(i, 3).value
                    row_data['data'] = sheet.cell(i, 4).value
                    row_data['http_method'] = sheet.cell(i, 5).value
                    row_data['expected_code'] = str(sheet.cell(i, 6).value)
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id + 1, 1).value
                    row_data['title'] = sheet.cell(case_id + 1, 2).value
                    row_data['url'] = sheet.cell(case_id + 1, 3).value
                    row_data['data'] = sheet.cell(case_id + 1, 4).value
                    row_data['http_method'] = sheet.cell(case_id + 1, 5).value
                    row_data['expected_code'] = str(sheet.cell(case_id + 1, 6).value)
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):  # 专门写回数据  i行号  result结果
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]  # 选择表单
        if sheet_name == "register":
            data = eval(sheet.cell(2, 4).value)  # 读取数据
            phone_num = data["mobilephone"]  # 获得手机号
            new_phone_num = int(phone_num) + 1
            data["mobilephone"] = new_phone_num  # 重新赋值
            sheet.cell(2, 4).value = str(data)  # 写回
            sheet.cell(i, 7).value = result
            sheet.cell(i, 8).value = TestResult
        else:
            sheet.cell(i, 7).value = result
            sheet.cell(i, 8).value = TestResult
        wb.save(file_name)  # 保存结果


if __name__ == '__main__':
    test_data = DoExcel.get_data(test_case_path)
    print(test_data)
